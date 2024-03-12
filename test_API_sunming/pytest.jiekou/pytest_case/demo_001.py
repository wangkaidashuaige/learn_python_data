from openpyxl import load_workbook
import json
import requests
import pytest
import allure

# 导入数据
wb = load_workbook("D:\pytest.jiekou\common\case.xlsx")
ws = wb["Worksheet"]
test_data = []
for x in range(2, len(tuple(ws.rows)) + 1):
    testcase_data = []
    for y in range(2, 7):
        testcase_data.append(ws.cell(row=x, column=y).value)
    test_data.append(testcase_data)


@allure.feature('我焯这个自动化测试谁写的 怎么这么牛逼')
class Test_api:
    # 全局变量
    token = 'e887fa892b9b38ddc38df9de1122c40a'



    # @allure.epic("一级目录")     #这个优先级贼TM的高！！！
    @allure.story("评价")
    @allure.title("标题")
    @allure.step
    # 设置用例优先级
    # @allure.severity("critical")
    @pytest.mark.parametrize("url,method,topic_data,code,msg", test_data)
    def test_create_topic(self, url, method, topic_data, code, msg):
        allure.dynamic.description('该说明为本次接口测试用例的用例说明：1：')
        headers = {"Authorization": Test_api.token}
        res = requests.post(url=url, json=json.loads(topic_data), headers=headers)
        print(res.json())
        assert res.status_code == 200
        assert 'msg' in res.json()
        assert res.json()['msg'] == code

    @allure.story("是谁这么牛逼")
    @allure.title("原来是我")
    def test_sunming(self):
        print("我焯这个自动化测试谁写的 怎么这么牛逼")
    @allure.story("步天罡，踏北斗")
    @allure.title("有秽皆除无妖不斩")
    @pytest.mark.parametrize("url,method,topic_data,code,msg", test_data)
    def test_003(self, url, method, topic_data, code, msg):
        allure.dynamic.description('该说明为本次接口测试用例的用例说明：1：')
        headers = {"Authorization": Test_api.token}
        res = requests.post(url=url, json=json.loads(topic_data), headers=headers)
        print(res.json())
        assert res.status_code == 200
        assert 'msg' in res.json()
        assert res.json()['msg'] == code
